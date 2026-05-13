from collections import Counter
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import BulkUpload, Complaint, User

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill("solid", fgColor="2F2A24")
SECTION_FILL = PatternFill("solid", fgColor="EDE7DF")
THIN_BORDER = Border(bottom=Side(style="thin", color="CFC5BA"))


def _dt(value):
    return value.replace(tzinfo=None) if value and getattr(value, "tzinfo", None) else value


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autosize(sheet, max_widths: dict[int, int] | None = None):
    max_widths = max_widths or {}
    for column_cells in sheet.columns:
        index = column_cells[0].column
        width = 12
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_widths.get(index, 42)))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _style_table(sheet):
    _style_header(sheet[1])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = THIN_BORDER


def _distribution(items: list[Complaint], attr: str) -> list[tuple[str, int]]:
    counter = Counter(str(getattr(item, attr) or "Not Analyzed") for item in items)
    return sorted(counter.items(), key=lambda row: (-row[1], row[0]))


def _write_complaints_sheet(wb: Workbook, complaints: list[Complaint]):
    sheet = wb.active
    sheet.title = "Analyzed Complaints"
    sheet.append([
        "Case ID",
        "Customer Name",
        "Customer Email",
        "Complaint Text",
        "Category",
        "Department",
        "Sentiment",
        "Priority",
        "Confidence Score",
        "Status",
        "Source",
        "Created At",
        "Analyzed At",
    ])
    for item in complaints:
        sheet.append([
            item.id,
            item.customer_name,
            item.customer_email or "",
            item.complaint_text,
            item.category or "",
            item.department or "",
            item.sentiment or "",
            item.priority or "",
            item.confidence_score if item.confidence_score is not None else "",
            item.status,
            item.source,
            _dt(item.created_at),
            _dt(item.analyzed_at),
        ])
    _style_table(sheet)
    sheet.column_dimensions["D"].width = 62
    for cell in sheet["D"][1:]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=13):
        for cell in row:
            cell.number_format = "yyyy-mm-dd hh:mm"
    _autosize(sheet, {4: 62})


def _write_section(sheet, start_row: int, title: str, rows: list[tuple[str, object]]) -> int:
    sheet.cell(start_row, 1, title)
    sheet.cell(start_row, 1).font = Font(bold=True, color="2F2A24")
    sheet.cell(start_row, 1).fill = SECTION_FILL
    sheet.cell(start_row, 2).fill = SECTION_FILL
    row = start_row + 1
    if not rows:
        rows = [("No data", "")]
    for label, value in rows:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, "" if value is None else value)
        row += 1
    return row + 1


def _write_summary_sheet(wb: Workbook, upload: BulkUpload, complaints: list[Complaint]):
    sheet = wb.create_sheet("Summary")
    generated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    confidence_values = [float(item.confidence_score) for item in complaints if item.confidence_score is not None]
    average_confidence = round(sum(confidence_values) / len(confidence_values), 2) if confidence_values else ""

    sheet["A1"] = "Bulk Upload Analysis Summary"
    sheet["A1"].font = Font(bold=True, size=16, color="2F2A24")
    sheet["A2"] = "Database-backed export for one upload batch"
    sheet["A2"].font = Font(color="666666")

    row = 4
    row = _write_section(sheet, row, "UPLOAD SUMMARY", [
        ("File name", upload.file_name),
        ("Upload ID", upload.id),
        ("Upload timestamp", _dt(upload.upload_timestamp)),
        ("Total rows", upload.total_rows),
        ("Processed rows", upload.processed_rows),
        ("Failed rows", upload.failed_rows),
        ("Average confidence", average_confidence),
        ("Export generated at", generated_at),
    ])
    row = _write_section(sheet, row, "CATEGORY DISTRIBUTION", _distribution(complaints, "category"))
    row = _write_section(sheet, row, "SENTIMENT DISTRIBUTION", _distribution(complaints, "sentiment"))
    _write_section(sheet, row, "PRIORITY BREAKDOWN", _distribution(complaints, "priority"))

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 34
    for rows in sheet.iter_rows():
        for cell in rows:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_upload_export(db: Session, user: User, upload_id: str) -> tuple[BytesIO, str]:
    upload = db.scalar(
        select(BulkUpload).where(
            BulkUpload.id == upload_id,
            BulkUpload.organization_id == user.organization_id,
        )
    )
    if not upload:
        raise LookupError("Upload not found")

    complaints = list(
        db.scalars(
            select(Complaint)
            .where(
                Complaint.organization_id == user.organization_id,
                Complaint.bulk_upload_id == upload.id,
            )
            .order_by(Complaint.created_at.asc())
        )
    )

    wb = Workbook()
    _write_complaints_sheet(wb, complaints)
    _write_summary_sheet(wb, upload, complaints)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, f"bulk_upload_{upload.id}_analysis.xlsx"
