from collections import Counter
from datetime import datetime, timezone
from io import BytesIO
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Complaint, User
from app.services import analytics
from app.services.complaints import ANALYSIS_FAILED, PENDING_ANALYSIS, SOLVED, get_filtered_complaints

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


HEADER_FILL = PatternFill("solid", fgColor="2F2A24")
SECTION_FILL = PatternFill("solid", fgColor="EDE7DF")
THIN_BORDER = Border(bottom=Side(style="thin", color="CFC5BA"))


def _dt(value):
    return value.replace(tzinfo=None) if value and getattr(value, "tzinfo", None) else value


def _is_analyzed(item: Complaint) -> bool:
    return item.status == SOLVED and item.confidence_score is not None


def _safe_sheet_value(value):
    if value is None:
        return ""
    return value


def _counter_rows(items: list[Complaint], attr: str) -> list[tuple[str, int]]:
    counter = Counter(str(getattr(item, attr) or "Not Analyzed") for item in items)
    return sorted(counter.items(), key=lambda row: (-row[1], row[0]))


def _all_workspace_complaints(db: Session, user: User) -> list[Complaint]:
    return list(
        db.scalars(
            select(Complaint)
            .where(Complaint.organization_id == user.organization_id)
            .order_by(Complaint.created_at.desc())
        )
    )


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autosize(sheet, max_widths: dict[int, int] | None = None):
    max_widths = max_widths or {}
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        width = 12
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            width = max(width, min(len(str(value)) + 2, max_widths.get(column_index, 42)))
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _style_table(sheet, header_row: int = 1):
    _style_header(sheet[header_row])
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = THIN_BORDER


def _write_complaints_sheet(wb: Workbook, title: str, complaints: list[Complaint]):
    sheet = wb.active
    sheet.title = title
    headers = [
        "Case ID",
        "Customer Name",
        "Customer Email",
        "Complaint Text",
        "Category",
        "Department",
        "Sentiment",
        "Priority",
        "Confidence Score",
        "Status",
        "Source",
        "Created At",
        "Analyzed At",
    ]
    sheet.append(headers)
    for item in complaints:
        analyzed = _is_analyzed(item)
        sheet.append([
            item.id,
            item.customer_name,
            item.customer_email,
            item.complaint_text,
            item.category if analyzed else "",
            item.department if analyzed else "",
            item.sentiment if analyzed else "",
            item.priority if analyzed else "",
            item.confidence_score if analyzed else "",
            item.status,
            item.source,
            _dt(item.created_at),
            _dt(item.analyzed_at) if analyzed else "",
        ])

    _style_table(sheet)
    sheet.column_dimensions["D"].width = 62
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=13):
        for cell in row:
            cell.number_format = "yyyy-mm-dd hh:mm"
    _autosize(sheet, {4: 62})


def _write_section(sheet, start_row: int, title: str, rows: list[tuple[str, object]]) -> int:
    sheet.cell(start_row, 1, title)
    sheet.cell(start_row, 1).font = Font(bold=True, color="2F2A24")
    sheet.cell(start_row, 1).fill = SECTION_FILL
    sheet.cell(start_row, 2).fill = SECTION_FILL
    row_index = start_row + 1
    if not rows:
        rows = [("No data", "")]
    for label, value in rows:
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, _safe_sheet_value(value))
        row_index += 1
    return row_index + 1


def _write_summary_sheet(wb: Workbook, db: Session, user: User, exported: list[Complaint], scope_label: str):
    sheet = wb.create_sheet("Summary")
    all_items = _all_workspace_complaints(db, user)
    solved_items = [item for item in all_items if item.status == SOLVED]
    confidence_values = [float(item.confidence_score) for item in exported if item.confidence_score is not None]
    generated_at = datetime.now(timezone.utc).replace(tzinfo=None)

    sheet["A1"] = f"{scope_label} Export Summary"
    sheet["A1"].font = Font(bold=True, size=16, color="2F2A24")
    sheet["A2"] = "Generated from PostgreSQL complaint records"
    sheet["A2"].font = Font(color="666666")

    row = 4
    row = _write_section(sheet, row, "GENERAL SUMMARY", [
        ("Exported rows", len(exported)),
        ("Total complaints", len(all_items)),
        ("Total solved complaints", len(solved_items)),
        ("Total pending analysis", sum(1 for item in all_items if item.status == PENDING_ANALYSIS)),
        ("Total failed analysis", sum(1 for item in all_items if item.status == ANALYSIS_FAILED)),
    ])
    row = _write_section(sheet, row, "CATEGORY BREAKDOWN", _counter_rows(exported, "category"))
    row = _write_section(sheet, row, "PRIORITY BREAKDOWN", _counter_rows(exported, "priority"))
    row = _write_section(sheet, row, "SENTIMENT BREAKDOWN", _counter_rows(exported, "sentiment"))
    row = _write_section(sheet, row, "DEPARTMENT BREAKDOWN", _counter_rows(exported, "department"))
    row = _write_section(sheet, row, "CONFIDENCE METRICS", [
        ("Average confidence score", round(mean(confidence_values), 2) if confidence_values else ""),
        ("Highest confidence", round(max(confidence_values), 2) if confidence_values else ""),
        ("Lowest confidence", round(min(confidence_values), 2) if confidence_values else ""),
    ])
    _write_section(sheet, row, "EXPORT INFO", [
        ("Export generated at", generated_at),
        ("Workspace/app name", user.organization_name or "Sentra AI"),
        ("Export scope", scope_label),
    ])

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 28
    for rows in sheet.iter_rows():
        for cell in rows:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_key_value_sheet(wb: Workbook, title: str, rows: list[tuple[str, object]]):
    sheet = wb.create_sheet(title)
    sheet.append(["Metric", "Value"])
    for label, value in rows:
        sheet.append([label, _safe_sheet_value(value)])
    _style_table(sheet)
    _autosize(sheet)


def _write_analytics_data_sheet(wb: Workbook, charts: dict):
    sheet = wb.create_sheet("Analytics Data")
    row_index = 1
    sections = [
        ("Monthly Complaint Volume", charts.get("monthly_complaint_volume", [])),
        ("Complaints by Category", charts.get("complaints_by_category", [])),
        ("Sentiment Trend", charts.get("sentiment_trend", [])),
        ("Resolution Time Trend", charts.get("resolution_time_trend", [])),
        ("Department Load", charts.get("department_load", [])),
        ("Source Mix", charts.get("source_mix", [])),
        ("Priority Distribution", charts.get("priority_distribution", [])),
    ]
    for section_name, records in sections:
        sheet.cell(row_index, 1, section_name)
        sheet.cell(row_index, 1).font = Font(bold=True, color="2F2A24")
        sheet.cell(row_index, 1).fill = SECTION_FILL
        row_index += 1
        keys = sorted({key for record in records for key in record.keys()})
        if not keys:
            sheet.cell(row_index, 1, "No data")
            row_index += 2
            continue
        for column_index, key in enumerate(keys, start=1):
            sheet.cell(row_index, column_index, key)
        _style_header(sheet[row_index])
        row_index += 1
        for record in records:
            for column_index, key in enumerate(keys, start=1):
                sheet.cell(row_index, column_index, _safe_sheet_value(record.get(key)))
            row_index += 1
        row_index += 2
    _autosize(sheet)


def build_complaint_export(db: Session, user: User, filters: dict, scope: str = "complaints") -> tuple[BytesIO, str]:
    normalized_scope = str(scope or "complaints").strip().lower()
    full_summary = normalized_scope in {"full", "full-summary", "full_summary"}
    effective_filters = {} if full_summary else dict(filters)
    exported = get_filtered_complaints(db, user, effective_filters)

    wb = Workbook()
    sheet_title = "All Complaints" if full_summary else "Solved Complaints"
    scope_label = "Full Summary" if full_summary else "Solved Complaints" if effective_filters.get("status") == SOLVED else "Complaints"
    _write_complaints_sheet(wb, sheet_title, exported)
    _write_summary_sheet(wb, db, user, exported, scope_label)

    if full_summary:
        summary = analytics.dashboard_summary(db, user)
        _write_key_value_sheet(wb, "Dashboard KPIs", [
            ("Total complaints", summary.get("total")),
            ("Pending analysis", summary.get("pending")),
            ("Analysis failed", summary.get("analysis_failed") or summary.get("in_progress")),
            ("Solved", summary.get("solved")),
            ("High priority", summary.get("high_priority")),
            ("Critical", summary.get("critical")),
            ("Average resolution hours", summary.get("avg_resolution_hours")),
            ("Average confidence", summary.get("avg_confidence")),
        ])
        _write_analytics_data_sheet(wb, analytics.analytics_charts(db, user))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    today = datetime.now(timezone.utc).strftime("%Y_%m_%d")
    prefix = "full_complaint_summary" if full_summary else "solved_complaints" if effective_filters.get("status") == SOLVED else "complaints"
    return output, f"{prefix}_{today}.xlsx"
