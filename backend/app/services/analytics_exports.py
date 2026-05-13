from collections import Counter
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import desc, select
from sqlalchemy.orm import Session

from app.models import BulkUpload, Complaint, User
from app.services import analytics
from app.services.complaints import ANALYSIS_FAILED, PENDING_ANALYSIS, SOLVED

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill("solid", fgColor="2F2A24")
SECTION_FILL = PatternFill("solid", fgColor="EDE7DF")
THIN_BORDER = Border(bottom=Side(style="thin", color="CFC5BA"))


def _dt(value):
    return value.replace(tzinfo=None) if value and getattr(value, "tzinfo", None) else value


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autosize(sheet, max_widths: dict[int, int] | None = None):
    max_widths = max_widths or {}
    for column_cells in sheet.columns:
        index = column_cells[0].column
        width = 12
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_widths.get(index, 42)))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _style_table(sheet):
    _style_header(sheet[1])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _percentage(count: int, total: int) -> float:
    return round((count / total) * 100, 2) if total else 0


def _distribution_rows(complaints: list[Complaint], attr: str) -> list[tuple[str, int, float]]:
    total = len(complaints)
    counter = Counter(str(getattr(item, attr) or "Not Analyzed") for item in complaints)
    return [(label, count, _percentage(count, total)) for label, count in sorted(counter.items(), key=lambda row: (-row[1], row[0]))]


def _write_distribution_sheet(wb: Workbook, title: str, label_header: str, rows: list[tuple[str, int, float]]):
    sheet = wb.create_sheet(title)
    sheet.append([label_header, "Count", "Percentage"])
    for label, count, percentage in rows:
        sheet.append([label, count, percentage])
    _style_table(sheet)
    for cell in sheet["C"][1:]:
        cell.number_format = "0.00"
    _autosize(sheet)


def _write_overview_sheet(wb: Workbook, db: Session, user: User, complaints: list[Complaint]):
    sheet = wb.active
    sheet.title = "Overview"
    summary = analytics.dashboard_summary(db, user)
    uploads = list(db.scalars(select(BulkUpload).where(BulkUpload.organization_id == user.organization_id)))
    generated_at = datetime.now(timezone.utc).replace(tzinfo=None)

    confidence_values = [float(item.confidence_score) for item in complaints if item.confidence_score is not None]
    sheet.append(["Metric", "Value"])
    rows = [
        ("Total complaints", summary.get("total", 0)),
        ("Solved complaints", summary.get("solved", 0)),
        ("Pending analysis", summary.get("pending", 0)),
        ("Failed analysis", summary.get("analysis_failed") or summary.get("in_progress", 0)),
        ("Average confidence", summary.get("avg_confidence", 0)),
        ("Highest confidence", round(max(confidence_values), 2) if confidence_values else ""),
        ("Lowest confidence", round(min(confidence_values), 2) if confidence_values else ""),
        ("Bulk upload batches", len(uploads)),
        ("Upload rows processed", sum(item.processed_rows for item in uploads)),
        ("Upload rows failed", sum(item.failed_rows for item in uploads)),
        ("Export generated at", generated_at),
        ("Workspace/app name", user.organization_name or "Sentra AI"),
    ]
    for row in rows:
        sheet.append(list(row))
    _style_table(sheet)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 28
    for cell in sheet["B"]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"


def _write_recent_sheet(wb: Workbook, complaints: list[Complaint]):
    sheet = wb.create_sheet("Recent Complaints")
    sheet.append([
        "Case ID",
        "Customer Name",
        "Complaint Text",
        "Category",
        "Sentiment",
        "Priority",
        "Status",
        "Confidence Score",
        "Created At",
        "Analyzed At",
    ])
    for item in complaints[:100]:
        analyzed = item.status == SOLVED and item.confidence_score is not None
        sheet.append([
            item.id,
            item.customer_name,
            item.complaint_text,
            item.category if analyzed else "",
            item.sentiment if analyzed else "",
            item.priority if analyzed else "",
            item.status,
            item.confidence_score if analyzed else "",
            _dt(item.created_at),
            _dt(item.analyzed_at) if analyzed else "",
        ])
    _style_table(sheet)
    sheet.column_dimensions["C"].width = 62
    for cell in sheet["C"][1:]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = "yyyy-mm-dd hh:mm"
    _autosize(sheet, {3: 62})


def _write_trends_sheet(wb: Workbook, charts: dict):
    trend_sources = [
        ("Monthly Complaint Volume", charts.get("monthly_complaint_volume", [])),
        ("Sentiment Trend", charts.get("sentiment_trend", [])),
        ("Resolution Time Trend", charts.get("resolution_time_trend", [])),
    ]
    if not any(records for _, records in trend_sources):
        return

    sheet = wb.create_sheet("Trends")
    row_index = 1
    for title, records in trend_sources:
        if not records:
            continue
        sheet.cell(row_index, 1, title)
        sheet.cell(row_index, 1).font = Font(bold=True, color="2F2A24")
        sheet.cell(row_index, 1).fill = SECTION_FILL
        row_index += 1
        keys = sorted({key for record in records for key in record.keys()})
        for column_index, key in enumerate(keys, start=1):
            sheet.cell(row_index, column_index, key)
        _style_header(sheet[row_index])
        row_index += 1
        for record in records:
            for column_index, key in enumerate(keys, start=1):
                sheet.cell(row_index, column_index, record.get(key, ""))
            row_index += 1
        row_index += 2
    _autosize(sheet)


def build_analytics_export(db: Session, user: User) -> tuple[BytesIO, str]:
    complaints = list(
        db.scalars(
            select(Complaint)
            .where(Complaint.organization_id == user.organization_id)
            .order_by(desc(Complaint.created_at))
        )
    )
    charts = analytics.analytics_charts(db, user)

    wb = Workbook()
    _write_overview_sheet(wb, db, user, complaints)
    _write_distribution_sheet(wb, "Category Distribution", "Category", _distribution_rows(complaints, "category"))
    _write_distribution_sheet(wb, "Sentiment Distribution", "Sentiment", _distribution_rows(complaints, "sentiment"))
    _write_distribution_sheet(wb, "Priority Distribution", "Priority", _distribution_rows(complaints, "priority"))
    _write_distribution_sheet(wb, "Department Distribution", "Department", _distribution_rows(complaints, "department"))
    _write_recent_sheet(wb, complaints)
    _write_trends_sheet(wb, charts)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    today = datetime.now(timezone.utc).strftime("%Y_%m_%d")
    return buffer, f"analytics_report_{today}.xlsx"
